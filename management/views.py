from django.shortcuts import render, redirect
from .models import Product, ScannedProducts, SoldProducts, CashierDynamicProducts, SoldProduct, \
    SoldProductHub, ScannedProductHeader, SoldOutProduct, ProductType, Notification, Expenses, Revenue, Income, \
    UserCreationValidation, DeviceInformation
from .forms import ProductForm, CreateUserForm, ScannedProductAddQuantityForm, ProductTypeForm, ChangeProductForm, \
    RegistrationValidationForm, ExpenseDetailsEditForm, AddExpenseForm
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import Group, User
from django.utils import timezone
from django.http import StreamingHttpResponse, JsonResponse, HttpResponseRedirect, HttpResponseBadRequest, \
    HttpResponse
from django.views.decorators import gzip
from django.db.models import Sum, Count
from datetime import datetime, timedelta
from django.contrib.humanize.templatetags.humanize import naturaltime
from datetime import date
from django.db.models import Q
from .decorators import unauthenticated_user, allowed_users, admin_group_required, has_expired, superuser
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from openpyxl import Workbook
import os
import subprocess
import getpass


@has_expired
@login_required(login_url='login')
@allowed_users(allowed_roles=['encoder', 'admin'])
def index(request):
    users = User.objects.all()
    product_form = ProductForm()
    product_type_form = ProductTypeForm()
    products = CashierDynamicProducts.objects.filter(cashier=1)
    # device_id = uuid.uuid4()  # Combines host ID and current time
    # print(device_id)

    current_date = timezone.now()
    if not Expenses.objects.filter(date_no_time=current_date).exists():
        Expenses.objects.get_or_create(date_no_time=current_date)

    admin_users = User.objects.filter(groups__name='admin')

    barcodes = []
    for product in products:
        barcodes.append(product.barcode)

    # cashier = User.objects.get(username=username)
    scanned_products = ScannedProducts.objects.filter(cashier=1)
    product_list = []
    for products in scanned_products:
        for product in products.product.all():
            products_data = [
                {'name': product.name,
                 'scanned_quantity': product.scanned_quantity,
                 'price': (product.scanned_quantity * product.price)},
            ]
            product_list.append(products_data)

    if request.method == 'POST':
        if 'add-new-product-btn' in request.POST:
            add_product_form = ProductForm(request.POST, request.FILES)
            if add_product_form.is_valid():
                product = add_product_form.save(commit=False)
                if Product.objects.filter(name=product.name).exists():
                    messages.error(request, "Product with that name already exists")
                elif Product.objects.filter(barcode=product.barcode).exists():
                    messages.error(request, "Product with that barcode already exists")
                else:
                    messages.success(request, "Product Successfully Added!")
                    product.save()

                    # Let's Create a dynamic products for cashiers
                    for user in users:
                        cashier_product = CashierDynamicProducts.objects.create(
                            cashier=user,
                            name=product.name,
                            type=product.type,
                            barcode=product.barcode,
                            price=product.price,
                            expiry_date=product.expiry_date,
                            quantity=product.quantity,
                            scanned_quantity=0,
                            image=product.image,
                        )
                        cashier_product.save()

        elif 'add-new-product-type-btn' in request.POST:
            product_type_form = ProductTypeForm(request.POST, request.FILES)

            if product_type_form.is_valid():
                product_type = product_type_form.save(commit=False)
                if ProductType.objects.filter(product_type=product_type.product_type).exists():
                    messages.error(request, "Product Type with that Designation Already Exists!")
                else:
                    product_type.save()
                    messages.success(request, f'New Product Type {product_type.product_type} Successfully Added')

                http_referrer = request.META.get('HTTP_REFERER')
                if http_referrer:
                    return HttpResponseRedirect(http_referrer)
                else:
                    return redirect('homepage')

    return render(request, 'base.html', {
        'product_form': product_form,
        'users': users,
        'scanned_products': product_list,
        'barcodes': barcodes,
        'product_type_form': product_type_form,
        # 'user_belongs_to_encoder': user_belongs_to_encoder,
        'admin_users': admin_users,
        # 'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
@admin_group_required
@login_required(login_url='login')
def dashboard(request):
    default_date = date.today()
    users = User.objects.all()
    notifications = Notification.objects.filter(removed=False)
    # product_type = ProductType.objects.all()

    current_date = timezone.now()
    thirty_days_ago = current_date - timedelta(days=30)

    expired_products = Product.objects.filter(expiry_date__lte=current_date, expiry_date__gte=thirty_days_ago)

    thirty_days_from_now = current_date + timedelta(days=30)

    upcoming_expiring_products = Product.objects.filter(expiry_date__gte=current_date,
                                                        expiry_date__lte=thirty_days_from_now)

    sold_products_objects = SoldProducts.objects.filter(date_sold_no_time__gte=thirty_days_ago)

    sold_products_date = {}
    for products in sold_products_objects:
        for product in products.product.all():
            product_date = products.date_sold_no_time
            if product_date not in sold_products_date:
                sold_products_date[product_date] = [product.sold_quantity]
            else:
                sold_products_date[product_date].append(product.sold_quantity)

    summed_sold_products_quantities = [sum(quantity_list) for product, quantity_list in sold_products_date.items()]
    product_sold_date = [product_date.strftime('%Y-%m-%d') for product_date, quantity_list in sold_products_date.items()]

    sold_products = SoldProduct.objects.filter(date_sold_no_time__gte=thirty_days_ago)

    sold_products_list = {}
    for products in sold_products:
        product = products.name
        if product not in sold_products_list:
            sold_products_list[product] = [products.sold_quantity]
        else:
            sold_products_list[product].append(products.sold_quantity)

    summed_products = {product: sum(quantity_list) for product, quantity_list in sold_products_list.items()}

    summed_products_quantities = [sum(quantity_list) for product, quantity_list in sold_products_list.items()]
    summed_products_products = [product for product, quantity_list in sold_products_list.items()]

    sold_products_revenue = SoldProducts.objects.filter(date_sold_no_time__gte=thirty_days_ago)
    # sold_products_revenue = SoldProducts.objects.all()
    sold_products_profit_revenue = {}
    for products in sold_products_revenue:
        product_date = products.date_sold_no_time
        if product_date not in sold_products_profit_revenue:
            sold_products_profit_revenue[product_date] = [products.total_price]
        else:
            sold_products_profit_revenue[product_date].append(products.total_price)

    summed_sold_product_revenue = [sum(product_price) for product_date, product_price in sold_products_profit_revenue.items()]
    sold_product_profit_date = [product_date.strftime('%Y-%m-%d') for product_date, product_price in sold_products_profit_revenue.items()]

    expenses_list = {}
    expenses = Expenses.objects.filter(date_no_time__gte=thirty_days_ago)
    # expenses = Expenses.objects.all()
    for expense in expenses:
        expense_date = expense.date_no_time
        if expense_date not in expenses_list:
            expenses_list[expense_date] = [expense.expense]
        else:
            expenses_list[expense_date].append(expense.expense)

    expense_sum = [sum(expense) for expense_date, expense in expenses_list.items()]
    expense_date = [expense_date.strftime('%Y-%m-%d') for expense_date, expense in expenses_list.items()]

    # for_income_date = [expense_date for expense_date, expense in expenses_list.items()]

    for product_date, product_price in sold_products_profit_revenue.items():
        revenue, created = Revenue.objects.get_or_create(date=product_date)
        revenue.revenue = sum(product_price)
        revenue.save()

    revenue_thirty_days = Revenue.objects.filter(date__gte=thirty_days_ago)

    rev_dates = []
    for rev in revenue_thirty_days:
        rev_dates.append(rev.date)
        for dater in rev_dates:
            derived_rev = Revenue.objects.get(date=dater)
            derived_exp = Expenses.objects.filter(date_no_time=dater)

            derived_inc, created = Income.objects.get_or_create(
                date=derived_rev.date,
                defaults={
                    'date_time': derived_rev.date_time
                }
            )

            derived_inc.income = derived_rev.revenue - sum([exp.expense for exp in derived_exp])
            derived_inc.save()

    incomes = Income.objects.filter(date__gte=thirty_days_ago).order_by('date')

    income_dict = {}
    for income in incomes:
        date_income = income.date
        if date_income not in income_dict:
            income_dict[date_income] = income.income
        else:
            income_dict[date_income].append(income.income)

    income_array = [inc for date_inc, inc in income_dict.items()]
    income_date = [date_inc.strftime('%Y-%m-%d') for date_inc, inc in income_dict.items()]
    # Filtration of graphs
    selected_month_sold_product = request.GET.get('selected-month')
    selected_type_sold_product = request.GET.get('type-filter')
    selected_cashier_sold_product = request.GET.get('cashier-branch-filter')

    sold_products_list_filtered = {}
    sold_products_date_filtered = {}
    sold_products_revenue_filtered = {}
    expense_list_filtered = {}
    income_array_filtered = {}

    if selected_month_sold_product and selected_cashier_sold_product and selected_type_sold_product:

        # Convert the selected date to a Python date object
        selected_date = datetime.strptime(selected_month_sold_product + '-01', '%Y-%m-%d').date()
        selected_year = selected_date.year
        selected_month = selected_date.month

        sold_products_filtered = SoldProduct.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            type__product_type=selected_type_sold_product,
            cashier__username=selected_cashier_sold_product,
        )

        for products in sold_products_filtered:
            product = products.name
            if product not in sold_products_list_filtered:
                sold_products_list_filtered[product] = [products.sold_quantity]
            else:
                sold_products_list_filtered[product].append(products.sold_quantity)

        sold_products_objects_filtered = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            product__type__product_type=selected_type_sold_product,
            cashier__username=selected_cashier_sold_product,
        )

        for products in sold_products_objects_filtered:
            for product in products.product.all():
                product_date = products.date_sold_no_time
                if product_date not in sold_products_date_filtered:
                    sold_products_date_filtered[product_date] = [product.sold_quantity]
                else:
                    sold_products_date_filtered[product_date].append(product.sold_quantity)

        sold_products_profits = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            product__type__product_type=selected_type_sold_product,
            cashier__username=selected_cashier_sold_product,
        )

        for products in sold_products_profits:
            product_date = products.date_sold_no_time
            for product in products.product.all():
                if product_date not in sold_products_revenue_filtered:
                    sold_products_revenue_filtered[product_date] = [product.price]
                else:
                    sold_products_revenue_filtered[product_date].append(product.price)

        expenses_filtered = Expenses.objects.filter(
            date_no_time__month=selected_month,
            date_no_time__year=selected_year,
        )
        # expenses = Expenses.objects.all()
        for expense in expenses_filtered:
            expense_date = expense.date_no_time
            if expense_date not in expense_list_filtered:
                expense_list_filtered[expense_date] = [expense.expense]
            else:
                expense_list_filtered[expense_date].append(expense.expense)

        incomes = Income.objects.filter(
            date__month=selected_month,
            date__year=selected_year,
        )
        for income in incomes:
            income_date = income.date
            if income_date not in income_array_filtered:
                income_array_filtered[income_date] = [income.income]
            else:
                income_array_filtered[income_date].append(income.income)

    if selected_month_sold_product and selected_cashier_sold_product == 'None' \
            and selected_type_sold_product == 'None':
        # Convert the selected date to a Python date object
        selected_date = datetime.strptime(selected_month_sold_product + '-01', '%Y-%m-%d').date()
        selected_year = selected_date.year
        selected_month = selected_date.month

        sold_products_filtered = SoldProduct.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
        )

        for products in sold_products_filtered:
            product = products.name
            if product not in sold_products_list_filtered:
                sold_products_list_filtered[product] = [products.sold_quantity]
            else:
                sold_products_list_filtered[product].append(products.sold_quantity)

        sold_products_objects_filtered = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
        )

        for products in sold_products_objects_filtered:
            for product in products.product.all():
                product_date = products.date_sold_no_time
                if product_date not in sold_products_date_filtered:
                    sold_products_date_filtered[product_date] = [product.sold_quantity]
                else:
                    sold_products_date_filtered[product_date].append(product.sold_quantity)

        sold_products_profits = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
        )

        for products in sold_products_profits:
            product_date = products.date_sold_no_time
            for product in products.product.all():
                if product_date not in sold_products_revenue_filtered:
                    sold_products_revenue_filtered[product_date] = [product.price]
                else:
                    sold_products_revenue_filtered[product_date].append(product.price)

        expenses_filtered = Expenses.objects.filter(
            date_no_time__month=selected_month,
            date_no_time__year=selected_year,
        )

        # expenses = Expenses.objects.all()
        for expense in expenses_filtered:
            expense_date = expense.date_no_time
            if expense_date not in expense_list_filtered:
                expense_list_filtered[expense_date] = [expense.expense]
            else:
                expense_list_filtered[expense_date].append(expense.expense)

        incomes = Income.objects.filter(
            date__month=selected_month,
            date__year=selected_year,
        )

        for income in incomes:
            income_date = income.date
            if income_date not in income_array_filtered:
                income_array_filtered[income_date] = [income.income]
            else:
                income_array_filtered[income_date].append(income.income)

    if selected_month_sold_product and selected_cashier_sold_product == 'None' \
            and selected_type_sold_product:

        # Convert the selected date to a Python date object
        selected_date = datetime.strptime(selected_month_sold_product + '-01', '%Y-%m-%d').date()
        selected_year = selected_date.year
        selected_month = selected_date.month

        sold_products_filtered = SoldProduct.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            type__product_type=selected_type_sold_product,
        )

        for products in sold_products_filtered:
            product = products.name
            if product not in sold_products_list_filtered:
                sold_products_list_filtered[product] = [products.sold_quantity]
            else:
                sold_products_list_filtered[product].append(products.sold_quantity)

        sold_products_objects_filtered = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            product__type__product_type=selected_type_sold_product,
        )

        for products in sold_products_objects_filtered:
            for product in products.product.all():
                product_date = products.date_sold_no_time
                if product_date not in sold_products_date_filtered:
                    sold_products_date_filtered[product_date] = [product.sold_quantity]
                else:
                    sold_products_date_filtered[product_date].append(product.sold_quantity)

        sold_products_profits = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            product__type__product_type=selected_type_sold_product,
        )

        for products in sold_products_profits:
            product_date = products.date_sold_no_time
            for product in products.product.all():
                if product_date not in sold_products_revenue_filtered:
                    sold_products_revenue_filtered[product_date] = [product.price]
                else:
                    sold_products_revenue_filtered[product_date].append(product.price)

        expenses_filtered = Expenses.objects.filter(
            date_no_time__month=selected_month,
            date_no_time__year=selected_year,
        )
        # expenses = Expenses.objects.all()
        for expense in expenses_filtered:
            expense_date = expense.date_no_time
            if expense_date not in expense_list_filtered:
                expense_list_filtered[expense_date] = [expense.expense]
            else:
                expense_list_filtered[expense_date].append(expense.expense)

        incomes = Income.objects.filter(
            date__month=selected_month,
            date__year=selected_year,
        )
        for income in incomes:
            income_date = income.date
            if income_date not in income_array_filtered:
                income_array_filtered[income_date] = [income.income]
            else:
                income_array_filtered[income_date].append(income.income)

    if selected_month_sold_product and selected_cashier_sold_product and selected_type_sold_product == 'None':

        # Convert the selected date to a Python date object
        selected_date = datetime.strptime(selected_month_sold_product + '-01', '%Y-%m-%d').date()
        selected_year = selected_date.year
        selected_month = selected_date.month

        sold_products_filtered = SoldProduct.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            cashier__username=selected_cashier_sold_product,
        )

        for products in sold_products_filtered:
            product = products.name
            if product not in sold_products_list_filtered:
                sold_products_list_filtered[product] = [products.sold_quantity]
            else:
                sold_products_list_filtered[product].append(products.sold_quantity)

        sold_products_objects_filtered = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            cashier__username=selected_cashier_sold_product,
        )

        for products in sold_products_objects_filtered:
            for product in products.product.all():
                product_date = products.date_sold_no_time
                if product_date not in sold_products_date_filtered:
                    sold_products_date_filtered[product_date] = [product.sold_quantity]
                else:
                    sold_products_date_filtered[product_date].append(product.sold_quantity)

        sold_products_profits = SoldProducts.objects.filter(
            date_sold_no_time__month=selected_month,
            date_sold_no_time__year=selected_year,
            cashier__username=selected_cashier_sold_product,
        )

        for products in sold_products_profits:
            product_date = products.date_sold_no_time
            for product in products.product.all():
                if product_date not in sold_products_revenue_filtered:
                    sold_products_revenue_filtered[product_date] = [product.price]
                else:
                    sold_products_revenue_filtered[product_date].append(product.price)

        expenses_filtered = Expenses.objects.filter(
            date_no_time__month=selected_month,
            date_no_time__year=selected_year,
        )
        # expenses = Expenses.objects.all()
        for expense in expenses_filtered:
            expense_date = expense.date_no_time
            if expense_date not in expense_list_filtered:
                expense_list_filtered[expense_date] = [expense.expense]
            else:
                expense_list_filtered[expense_date].append(expense.expense)

        incomes = Income.objects.filter(
            date__month=selected_month,
            date__year=selected_year,
        )
        for income in incomes:
            income_date = income.date
            if income_date not in income_array_filtered:
                income_array_filtered[income_date] = [income.income]
            else:
                income_array_filtered[income_date].append(income.income)

    # Filtered Products Quantity
    summed_sold_products_quantities_filtered = [
        sum(quantity_list) for product, quantity_list in sold_products_date_filtered.items()
    ]
    product_sold_date_filtered = [
        product_date.strftime('%Y-%m-%d') for product_date, quantity_list in sold_products_date_filtered.items()
    ]
    # Filtered Product Sales
    summed_products_quantities_filtered = [
        sum(quantity_list) for product, quantity_list in sold_products_list_filtered.items()
    ]
    summed_products_products_filtered = [
        product for product, quantity_list in sold_products_list_filtered.items()
    ]
    # Filtered Daily Profits
    summed_sold_product_revenue_filtered = [
        sum(product_price) for product_date, product_price in sold_products_revenue_filtered.items()
    ]
    sold_product_revenue_date_filtered = [
        product_date.strftime('%Y-%m-%d') for product_date, product_price in sold_products_revenue_filtered.items()
    ]

    expense_sum_filtered = [sum(expense) for expense_date, expense in expense_list_filtered.items()]
    expense_date_filtered = [expense_date.strftime('%Y-%m-%d') for expense_date, expense in expense_list_filtered.items()]

    daily_income_filtered = [sum(income) for date, income in income_array_filtered.items()]
    daily_income_date_filtered = [date.strftime('%Y-%m-%d') for date, income in income_array_filtered.items()]

    revenue_list = {}
    revenues = SoldProducts.objects.all()
    for revenue in revenues:
        rev = revenue.total_price
        if rev not in revenue_list:
            revenue_list[rev] = [rev]
        else:
            revenue_list[rev].append(rev)

    total_revenue = [sum(rev_array) for rev, rev_array in revenue_list.items()]
    total_revenue_sum = sum(total_revenue)

    return render(request, 'dashboard.html', {
        'income_date': income_date,
        'income_array': income_array,
        'total_revenue': total_revenue,
        'total_revenue_sum': total_revenue_sum,
        'summed_products': summed_products,
        'summed_products_quantities': summed_products_quantities,
        'summed_products_products': summed_products_products,
        'sold_products_date': sold_products_date,
        'sold_products_list': sold_products_list,
        'summed_sold_products_quantities': summed_sold_products_quantities,
        'product_sold_date': product_sold_date,
        'expired_products': expired_products,
        'upcoming_expiring_products': upcoming_expiring_products,
        'notifications': notifications,
        'default_date': default_date,
        # 'product_type': product_type,
        # 'product_types': prod_types,
        'users': users,
        # Expenses Data
        'expense_sum': expense_sum,
        'expense_date': expense_date,
        # Sold Product Profit
        'summed_sold_product_revenue': summed_sold_product_revenue,
        'sold_product_profit_date': sold_product_profit_date,
        # Sold Product Filters
        'selected_month_sold_product': selected_month_sold_product,
        'selected_type_sold_product': selected_type_sold_product,
        'selected_cashier_sold_product': selected_cashier_sold_product,
        # Filtered Dashboard Data Sold Products
        'summed_products_quantities_filtered': summed_products_quantities_filtered,
        'summed_products_products_filtered': summed_products_products_filtered,
        # Filtered Dashboard Data Daily Sales
        'summed_sold_products_quantities_filtered': summed_sold_products_quantities_filtered,
        'product_sold_date_filtered': product_sold_date_filtered,
        # Filtered Dashboard Data Daily Profits
        'summed_sold_product_revenue_filtered': summed_sold_product_revenue_filtered,
        'sold_product_revenue_date_filtered': sold_product_revenue_date_filtered,
        # Filtered Dashboard Data Expenses
        'expense_sum_filtered': expense_sum_filtered,
        'expense_date_filtered': expense_date_filtered,
        # Filtered Income
        'income_array_filtered': daily_income_filtered,
        'for_income_date_filtered': daily_income_date_filtered,
    })


@has_expired
def expenses_page(request):
    expenses = Expenses.objects.all().order_by('-date_no_time')

    expenses_array = expenses.values('expense')
    total_expenses = expenses_array.aggregate(total_exp=Sum('expense'))['total_exp']

    selected_date_str = request.GET.get('expenses-selected-date')

    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()

    if selected_date_str:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m').date()
    else:
        selected_date = None

    if selected_date:
        filtered_expenses = Expenses.objects.filter(
            date_no_time__month=selected_date.month,
            date_no_time__year=selected_date.year
        ).order_by('-date_no_time')

        filtered_expenses_array = filtered_expenses.values('expense')
        filtered_expenses_sum = filtered_expenses_array.aggregate(total_filt_exp=Sum('expense'))['total_filt_exp']
    else:
        filtered_expenses = []
        filtered_expenses_sum = ''

    return render(request, 'expenses_page.html', {
        'expenses': expenses,
        'filtered_expenses': filtered_expenses,
        'selected_date': selected_date,
        'total_expenses': total_expenses,
        'filtered_expenses_sum': filtered_expenses_sum,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
def download_expense_excel(request):
    # Query the Income model to get the data
    expenses = Expenses.objects.all().order_by('-date_no_time')

    expenses_array = expenses.values('expense')
    total_expenses = expenses_array.aggregate(total_exp=Sum('expense'))['total_exp']

    total_expenses_row = ['Total Expenses', total_expenses]

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ['Amount', 'Date']
    ws.append(headers)

    # Add data to the worksheet
    for expense in expenses:
        row_data = [expense.expense, expense.date_no_time.strftime('%Y-%m-%d')]
        ws.append(row_data)

    ws.append(total_expenses_row)

    # Create a response object with the appropriate content type for Excel files
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expenses_data.xlsx'

    # Save the workbook to the response object
    wb.save(response)

    return response


@has_expired
def download_expense_excel_filtered(request):
    # Query the Income model to get the data
    selected_date_str = request.GET.get('expenses-selected-date')

    if selected_date_str:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m').date()
    else:
        selected_date = None

    if selected_date:
        filtered_expenses = Expenses.objects.filter(
            date_no_time__month=selected_date.month,
            date_no_time__year=selected_date.year
        ).order_by('-date_no_time')

        filtered_expenses_array = filtered_expenses.values('expense')
        filtered_expenses_sum = filtered_expenses_array.aggregate(total_filt_exp=Sum('expense'))['total_filt_exp']
    else:
        filtered_expenses = []
        filtered_expenses_sum = ''
    total_expenses_row = ['Total Expenses', filtered_expenses_sum]
    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ['Amount', 'Date']
    ws.append(headers)

    # Add data to the worksheet
    for expense in filtered_expenses:
        row_data = [expense.expense, expense.date_no_time.strftime('%Y-%m-%d')]
        ws.append(row_data)

    ws.append(total_expenses_row)

    # Create a response object with the appropriate content type for Excel files
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=expenses_data-{selected_date}.xlsx'

    # Save the workbook to the response object
    wb.save(response)

    return response


@has_expired
def modify_expense_details(request, id, expense_date):
    formatted_date = datetime.strptime(expense_date, '%Y-%m-%d')
    expense = Expenses.objects.get(id=id, date_no_time=formatted_date)
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    if request.method == 'POST':
        form = ExpenseDetailsEditForm(request.POST, request.FILES, instance=expense)
        expense_details = form.save(commit=False)
        expense.expense = expense_details.expense
        expense.description = expense_details.description
        expense.image = expense_details.image
        expense.save()

    expense_form = ExpenseDetailsEditForm(instance=expense)
    return render(request, 'expense_detail.html', {
        'expense': expense,
        'expense_form': expense_form,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
def income_page(request):
    incomes = Income.objects.all().order_by('-date')

    income_array = incomes.values('income')
    total_income = income_array.aggregate(total_sum=Sum('income'))['total_sum']

    selected_month_string = request.GET.get('income-page-selected-date')
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    if selected_month_string:
        selected_date = datetime.strptime(selected_month_string, '%Y-%m').date()
    else:
        selected_date = None

    if selected_date:
        filtered_income = Income.objects.filter(
            date__month=selected_date.month,
            date__year=selected_date.year).order_by('-date')

        filtered_income_array = filtered_income.values('income')
        filtered_income_sum = filtered_income_array.aggregate(filtered_sum=Sum('income'))['filtered_sum']
    else:
        filtered_income = []
        filtered_income_sum = ''

    return render(request, 'income_page.html', {
        'incomes': incomes,
        'filtered_income': filtered_income,
        'selected_date': selected_date,
        'total_income': total_income,
        'filtered_income_sum': filtered_income_sum,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
def download_income_excel(request):
    # Query the Income model to get the data
    incomes = Income.objects.all().order_by('-date')

    income_array = incomes.values('income')
    total_income = income_array.aggregate(total_sum=Sum('income'))['total_sum']

    total_income_row = ['Total Income', total_income]

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ['Income', 'Date']
    ws.append(headers)

    # Add data to the worksheet
    for income in incomes:
        row_data = [income.income, income.date.strftime('%Y-%m-%d')]
        ws.append(row_data)

    ws.append(total_income_row)

    # Create a response object with the appropriate content type for Excel files
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=income_data.xlsx'

    # Save the workbook to the response object
    wb.save(response)

    return response


@has_expired
def download_income_excel_filtered(request):
    # Query the Income model to get the data
    selected_month_string = request.GET.get('income-page-selected-date')

    if selected_month_string:
        selected_date = datetime.strptime(selected_month_string, '%Y-%m').date()
    else:
        selected_date = None

    if selected_date:
        filtered_income = Income.objects.filter(
            date__month=selected_date.month,
            date__year=selected_date.year).order_by('-date')

        filtered_income_array = filtered_income.values('income')
        filtered_income_sum = filtered_income_array.aggregate(filtered_sum=Sum('income'))['filtered_sum']

    else:
        filtered_income = []
        filtered_income_sum = ''

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ['Income', 'Date']
    ws.append(headers)

    # Add data to the worksheet
    for income in filtered_income:
        row_data = [income.income, income.date.strftime('%Y-%m-%d')]
        ws.append(row_data)

    ws.append(['Total Income', filtered_income_sum])

    # Create a response object with the appropriate content type for Excel files
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=income_data-{selected_date}.xlsx'

    # Save the workbook to the response object
    wb.save(response)

    return response


@has_expired
def revenue_page(request):
    revenues = Revenue.objects.all()
    selected_month_string = request.GET.get('revenue-page-selected-month')

    revenue_array = revenues.values('revenue')
    revenue_sum = revenue_array.aggregate(total_rev=Sum('revenue'))['total_rev']
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    if selected_month_string:
        selected_month = datetime.strptime(selected_month_string, '%Y-%m').date()
    else:
        selected_month = None

    if selected_month:
        filtered_revenues = Revenue.objects.filter(
            date__month=selected_month.month,
            date__year=selected_month.year).order_by('-date')

        filtered_revenue_array = filtered_revenues.values('revenue')
        filtered_revenue_sum = filtered_revenue_array.aggregate(
            total_rev_filtered=Sum('revenue'))['total_rev_filtered']
    else:
        filtered_revenues = []
        filtered_revenue_sum= ''

    return render(request, 'revenue_page.html', {
        'revenues': revenues,
        'selected_month': selected_month,
        'filtered_revenues': filtered_revenues,
        'revenue_sum': revenue_sum,
        'filtered_revenue_sum': filtered_revenue_sum,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
def download_revenue_excel(request):
    # Query the Income model to get the data
    revenues = Revenue.objects.all().order_by('-date')

    revenue_array = revenues.values('revenue')
    revenue_sum = revenue_array.aggregate(total_rev=Sum('revenue'))['total_rev']

    total_revenue_row = ['Total Revenue', revenue_sum]

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ['Amount', 'Date']
    ws.append(headers)

    # Add data to the worksheet
    for revenue in revenues:
        row_data = [revenue.revenue, revenue.date.strftime('%Y-%m-%d')]
        ws.append(row_data)

    ws.append(total_revenue_row)

    # Create a response object with the appropriate content type for Excel files
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=revenue_data.xlsx'

    # Save the workbook to the response object
    wb.save(response)

    return response


@has_expired
def download_revenue_excel_filtered(request):
    # Query the Income model to get the data
    selected_month_string = request.GET.get('revenue-page-selected-month')

    if selected_month_string:
        selected_month = datetime.strptime(selected_month_string, '%Y-%m').date()
    else:
        selected_month = None

    if selected_month:
        filtered_revenues = Revenue.objects.filter(
            date__month=selected_month.month,
            date__year=selected_month.year).order_by('-date')

        filtered_revenue_array = filtered_revenues.values('revenue')
        filtered_revenue_sum = filtered_revenue_array.aggregate(
            total_rev_filtered=Sum('revenue'))['total_rev_filtered']

    else:
        filtered_revenues = []
        filtered_revenue_sum = ''

    total_revenue_row = ['Total Revenue', filtered_revenue_sum]

    # Create a new Excel workbook and add a worksheet
    wb = Workbook()
    ws = wb.active

    # Add headers to the worksheet
    headers = ['Amount', 'Date']
    ws.append(headers)

    # Add data to the worksheet
    for revenue in filtered_revenues:
        row_data = [revenue.revenue, revenue.date.strftime('%Y-%m-%d')]
        ws.append(row_data)

    ws.append(total_revenue_row)

    # Create a response object with the appropriate content type for Excel files
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=revenue_data{selected_month}.xlsx'

    # Save the workbook to the response object
    wb.save(response)

    return response


@has_expired
@unauthenticated_user
def authentication(request):
    notifications = Notification.objects.filter(removed=False)
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            if user.groups.filter(name='cashier').exists():
                return redirect('scanned-products-cashier', username=username)
            else:
                return redirect('homepage')
        else:
            messages.error(request, "Invalid Form Data")

    return render(request, 'login.html', {

    })


@has_expired
@unauthenticated_user
def registration_validation(request):
    if request.method == 'POST':
        form = RegistrationValidationForm(request.POST)
        if form.is_valid():

            user = form.save(commit=False)
            password1 = form.cleaned_data['password1']
            password2 = form.cleaned_data['password2']

            if User.objects.filter(email=user.email).exists() or\
                    UserCreationValidation.objects.filter(email=user.email).exists():
                messages.error(request, 'Email Already Exists')
            if password1 == password2 and not User.objects.filter(email=user.email).exists() and not \
                    UserCreationValidation.objects.filter(email=user.email).exists():
                hashed_password = make_password(password1)
                user.password = hashed_password

                user.save()

                admins = User.objects.filter(groups__name='admin')
                for admin in admins:
                    notification, created = Notification.objects.get_or_create(
                        user=admin,
                        title=f'Registrant {user.username} awaiting verification',
                        message=f'Registrant {user.username} registered on {user.date_time} is '
                                f'awaiting to be verified ot denied',
                        identifier=f'Identifier {user.username}, {user.id}',
                    )
                    notification.save()

                # Create Register Instance
                # registration_instance = Registrations.objects.create(instance=user.username)
                # registration_instance.save()

                rendered_template = render(request, 'registration_response.html', {'variable': 'value'})

                # Create an HttpResponse with the rendered template as content
                response = HttpResponse(rendered_template)
                return response

            elif password1 != password2:
                messages.error(request, 'Password not the same')

            if User.objects.filter(username=user.username).exists() or \
                    UserCreationValidation.objects.filter(username=user.username).exists():
                messages.error(request, 'Username Already Exists')

    else:
        form = RegistrationValidationForm()
    return render(request, 'registration_validation.html', {
        'form': form,
    })


@has_expired
@admin_group_required
@login_required(login_url='login')
def validation_registrants(request):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    registrants = UserCreationValidation.objects.all()
    return render(request, 'confirming_registration.html', {
        'registrants': registrants,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
@unauthenticated_user
def register(request):
    notifications = Notification.objects.all()
    if request.user.is_authenticated:
        return redirect('homepage')
    else:
        form = CreateUserForm
        if request.method == 'POST':
            form = CreateUserForm(request.POST)
            if form.is_valid():
                user_created = form.save(commit=False)
                user_created.save()

                new_user = User.objects.get(username=user_created.username)

                # Create a ScannedProducts Objects to address the scalability concern
                cashier, created = ScannedProducts.objects.get_or_create(
                    cashier=new_user,
                    summed_product_price=0,
                )
                cashier.save()

                scanned_product_header, created = ScannedProductHeader.objects.get_or_create(
                    cashier=new_user,
                )
                scanned_product_header.save()
                # Let's Create a Product for the Registered Cashier so as not to tamper the original product's details
                products = Product.objects.all()
                for product in products:
                    dynamic_cashier_products = CashierDynamicProducts.objects.create(cashier=user_created,
                                                                                     name=product.name,
                                                                                     type=product.type,
                                                                                     barcode=product.barcode,
                                                                                     price=product.price,
                                                                                     expiry_date=product.expiry_date,
                                                                                     quantity=product.quantity,
                                                                                     scanned_quantity=0,
                                                                                     image=product.image,
                                                                                     )
                    dynamic_cashier_products.save()

                # group = Group.objects.get(name='users')
                # userForm.groups.add(group)

                messages.success(request, 'Account successfully created ' + user_created.username)
                return redirect('login')

    return render(request, 'register.html', {'form': form})


@has_expired
@admin_group_required
@login_required(login_url='login')
def confirm_registration(request, username):
    registrant = UserCreationValidation.objects.get(username=username)
    user_created = User.objects.create(
        username=registrant.username,
        password=registrant.password,
        email=registrant.email
    )

    group = Group.objects.get(id=registrant.group.id)
    user_created.groups.add(group)
    user_created.save()
    registrant.delete()

    new_user = User.objects.get(username=user_created.username)

    if new_user.groups.filter(name='cashier').exists():
        # Create a ScannedProducts Objects to address the scalability concern
        cashier, created = ScannedProducts.objects.get_or_create(
            cashier=new_user,
            summed_product_price=0,
        )
        cashier.save()

        scanned_product_header, created = ScannedProductHeader.objects.get_or_create(
            cashier=new_user,
        )
        scanned_product_header.save()
        # Let's Create a Product for the Registered Cashier so as not to tamper the original product's details
        products = Product.objects.all()
        for product in products:
            dynamic_cashier_products = CashierDynamicProducts.objects.create(cashier=user_created,
                                                                             name=product.name,
                                                                             type=product.type,
                                                                             barcode=product.barcode,
                                                                             price=product.price,
                                                                             expiry_date=product.expiry_date,
                                                                             quantity=product.quantity,
                                                                             scanned_quantity=0,
                                                                             image=product.image,
                                                                             )
            dynamic_cashier_products.save()

        # group = Group.objects.get(name='users')
        # userForm.groups.add(group)

        messages.success(request, 'Account successfully created ' + user_created.username)
    messages.success(request, 'Account successfully created ' + user_created.username)
    return redirect('login')


@has_expired
@admin_group_required
@login_required(login_url='login')
def deny_registration(request, id):
    registrant = UserCreationValidation.objects.get(id=id)
    registrant.delete()

    http_referrer = request.META.get('HTTP_REFERER')
    if http_referrer:
        return HttpResponseRedirect(http_referrer)
    else:
        return redirect('homepage')


def logout_user(request):
    logout(request)
    return redirect('login')


@has_expired
@allowed_users(allowed_roles=['admin'])
@login_required(login_url='login')
def scanning_products(request, username):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    notifications = Notification.objects.filter(removed=False)
    user = User.objects.get(username=username)
    cashier = User.objects.get(username=username)
    scanned_products = ScannedProducts.objects.filter(cashier=user)
    products = CashierDynamicProducts.objects.filter(cashier=user, quantity__gt=0)
    scanned_product_header, created = ScannedProductHeader.objects.get_or_create(cashier=user)
    scanned_product_header.save()
    scanned_product_list, created = ScannedProducts.objects.get_or_create(cashier=user)
    scanned_products_header = ScannedProductHeader.objects.get(cashier=user)

    add_quantity_form = ScannedProductAddQuantityForm()

    searched_product_name_input = request.GET.get('search-product-input')
    searched_products = None
    if searched_product_name_input:
        searched_products = CashierDynamicProducts.objects.filter(
            cashier=user,
            quantity__gt=0,
            name__icontains=searched_product_name_input
        )

    scanned_length = ScannedProducts.objects.filter(cashier=cashier).annotate(scanned_length=Count('product'))
    scanned_sum = scanned_length.aggregate(length_sum=Sum('scanned_length'))['length_sum']

    scanned_barcode = request.POST.get('scanned-barcode')
    if request.method == 'POST':
        if CashierDynamicProducts.objects.filter(cashier=user, barcode=scanned_barcode, quantity__gt=0).exists():
            scanned_product = CashierDynamicProducts.objects.get(barcode=scanned_barcode, cashier=user)
            scanned_product_list.product.add(scanned_product)

            scanned_product_list.summed_product_price += scanned_product.price
            scanned_product_list.save()

            scanned_product.scanned_quantity += 1
            scanned_product.quantity -= 1
            scanned_product.save()

            main_products = Product.objects.filter(name=scanned_product.name)
            for main_product in main_products:
                main_product.quantity = scanned_product.quantity
                main_product.save()

            dynamic_products = CashierDynamicProducts.objects.filter(name=scanned_product.name)
            for dynamic_product in dynamic_products:
                dynamic_product.quantity = scanned_product.quantity
                dynamic_product.save()

            # Create scanned header
            scanned_product_header, created = ScannedProductHeader.objects.get_or_create(cashier=user)
            scanned_product_header.name = scanned_product.name
            scanned_product_header.barcode = scanned_product.barcode
            scanned_product_header.expiry_date = scanned_product.expiry_date
            scanned_product_header.price = scanned_product.price
            scanned_product_header.save()

            referring_url = request.META.get('HTTP_REFERER')

            if referring_url:
                # Redirect back to the referring page
                return HttpResponseRedirect(referring_url)
            else:
                # If there's no referring URL, redirect to a default page
                return redirect('scanned-products', username=user.username)

        elif 'quantity-configuration-form' in request.POST:
            product_quantity_id = request.POST.get('product-quantity-id')
            product_quantity = int(request.POST.get('product-quantity-quantity'))
            product_quantity_cashier = request.POST.get('product-quantity-cashier')

            scanned_product = CashierDynamicProducts.objects.get(
                id=product_quantity_id,
                cashier=product_quantity_cashier,
            )

            if product_quantity < scanned_product.scanned_quantity:
                difference = scanned_product.scanned_quantity - product_quantity
                scanned_product.quantity += difference
                scanned_product.scanned_quantity = product_quantity
                # scanned_product.quantity -= product_quantity
                scanned_product.save()
            elif scanned_product.scanned_quantity < product_quantity:
                difference = product_quantity - scanned_product.scanned_quantity
                scanned_product.quantity -= difference
                scanned_product.scanned_quantity = product_quantity
                scanned_product.save()

            main_product = Product.objects.get(
                name=scanned_product.name
            )
            main_product.quantity = scanned_product.quantity
            main_product.save()

            dynamic_products = CashierDynamicProducts.objects.filter(
                name=scanned_product.name
            ).exclude(cashier=product_quantity_cashier)

            for dy_prod in dynamic_products:
                dy_prod.quantity = scanned_product.quantity
                dy_prod.save()

            scanned_product_cashier_sum = []

            for hark in scanned_product_list.product.all():
                scanned_product_cashier_sum.append(int(hark.multiply()))

            scanned_product_list.summed_product_price = sum(scanned_product_cashier_sum)
            scanned_product_list.save()

            referring_url = request.META.get('HTTP_REFERER')

            if referring_url:
                # Redirect back to the referring page
                return HttpResponseRedirect(referring_url)
            else:
                # If there's no referring URL, redirect to a default page
                return redirect('scanned-products', username=user.username)

        elif 'searched-products-btn' in request.POST:
            searched_product_ids = request.POST.getlist('searched-product-ids')
            for product_id in searched_product_ids:
                searched_product = CashierDynamicProducts.objects.get(id=int(product_id))
                scanned_product_list.product.add(searched_product)

                scanned_product_list.summed_product_price += searched_product.price
                scanned_product_list.save()

                searched_product.scanned_quantity += 1
                searched_product.quantity -= 1
                searched_product.save()

                main_product = Product.objects.get(
                    name=searched_product.name
                )
                main_product.quantity = searched_product.quantity
                main_product.save()

                dynamic_products = CashierDynamicProducts.objects.filter(
                    name=searched_product.name
                ).exclude(cashier=user)

                for dy_prod in dynamic_products:
                    dy_prod.quantity = searched_product.quantity
                    dy_prod.save()

            referring_url = request.META.get('HTTP_REFERER')

            if referring_url:
                # Redirect back to the referring page
                return HttpResponseRedirect(referring_url)
            else:
                # If there's no referring URL, redirect to a default page
                return redirect('scanned-products', username=user.username)

    return render(request, 'scanned_product.html', {
        'scanned_sum': scanned_sum,
        'scanned_products': scanned_products,
        'products': products,
        'user': user,
        'scanned_product_header': scanned_product_header,
        'add_quantity_form': add_quantity_form,
        'notifications': notifications,
        'user_belongs_to_admin': user_belongs_to_admin,
        'scanned_products_header': scanned_products_header,
        'searched_products': searched_products,
        'searched_product_name_input': searched_product_name_input,
    })


@has_expired
@allowed_users(allowed_roles=['cashier', 'admin'])
@login_required(login_url='login')
def scanning_products_cashier(request, username):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    notifications = Notification.objects.filter(removed=False)
    user = User.objects.get(username=username)
    cashier = User.objects.get(username=username)
    scanned_products = ScannedProducts.objects.filter(cashier=user)
    products = CashierDynamicProducts.objects.filter(cashier=user, quantity__gt=0)
    scanned_product_header, created = ScannedProductHeader.objects.get_or_create(cashier=user)
    scanned_product_header.save()
    scanned_product_list, created = ScannedProducts.objects.get_or_create(cashier=user)
    scanned_products_header = ScannedProductHeader.objects.get(cashier=user)

    add_quantity_form = ScannedProductAddQuantityForm()

    searched_product_name_input = request.GET.get('search-product-input')
    searched_products = None
    if searched_product_name_input:
        searched_products = CashierDynamicProducts.objects.filter(
            cashier=user,
            quantity__gt=0,
            name__icontains=searched_product_name_input
        )

    scanned_length = ScannedProducts.objects.filter(cashier=cashier).annotate(scanned_length=Count('product'))
    scanned_sum = scanned_length.aggregate(length_sum=Sum('scanned_length'))['length_sum']

    scanned_barcode = request.POST.get('scanned-barcode')
    if request.method == 'POST':
        if CashierDynamicProducts.objects.filter(cashier=user, barcode=scanned_barcode, quantity__gt=0).exists():
            scanned_product = CashierDynamicProducts.objects.get(barcode=scanned_barcode, cashier=user)
            scanned_product_list.product.add(scanned_product)

            scanned_product_list.summed_product_price += scanned_product.price
            scanned_product_list.save()

            scanned_product.scanned_quantity += 1
            scanned_product.quantity -= 1
            scanned_product.save()

            main_products = Product.objects.filter(name=scanned_product.name)
            for main_product in main_products:
                main_product.quantity = scanned_product.quantity
                main_product.save()

            dynamic_products = CashierDynamicProducts.objects.filter(name=scanned_product.name)
            for dynamic_product in dynamic_products:
                dynamic_product.quantity = scanned_product.quantity
                dynamic_product.save()

            # Create scanned header
            scanned_product_header, created = ScannedProductHeader.objects.get_or_create(cashier=user)
            scanned_product_header.name = scanned_product.name
            scanned_product_header.barcode = scanned_product.barcode
            scanned_product_header.expiry_date = scanned_product.expiry_date
            scanned_product_header.price = scanned_product.price
            scanned_product_header.save()

            referring_url = request.META.get('HTTP_REFERER')

            if referring_url:
                # Redirect back to the referring page
                return HttpResponseRedirect(referring_url)
            else:
                # If there's no referring URL, redirect to a default page
                return redirect('scanned-products', username=user.username)

        elif 'quantity-configuration-form' in request.POST:
            product_quantity_id = request.POST.get('product-quantity-id')
            product_quantity = int(request.POST.get('product-quantity-quantity'))
            product_quantity_cashier = request.POST.get('product-quantity-cashier')

            scanned_product = CashierDynamicProducts.objects.get(
                id=product_quantity_id,
                cashier=product_quantity_cashier,
            )

            if product_quantity < scanned_product.scanned_quantity:
                difference = scanned_product.scanned_quantity - product_quantity
                scanned_product.quantity += difference
                scanned_product.scanned_quantity = product_quantity
                # scanned_product.quantity -= product_quantity
                scanned_product.save()
            elif scanned_product.scanned_quantity < product_quantity:
                difference = product_quantity - scanned_product.scanned_quantity
                scanned_product.quantity -= difference
                scanned_product.scanned_quantity = product_quantity
                scanned_product.save()

            main_product = Product.objects.get(
                name=scanned_product.name
            )
            main_product.quantity = scanned_product.quantity
            main_product.save()

            dynamic_products = CashierDynamicProducts.objects.filter(
                name=scanned_product.name
            ).exclude(cashier=product_quantity_cashier)

            for dy_prod in dynamic_products:
                dy_prod.quantity = scanned_product.quantity
                dy_prod.save()

            scanned_product_cashier_sum = []

            for hark in scanned_product_list.product.all():
                scanned_product_cashier_sum.append(int(hark.multiply()))

            scanned_product_list.summed_product_price = sum(scanned_product_cashier_sum)
            scanned_product_list.save()

            referring_url = request.META.get('HTTP_REFERER')

            if referring_url:
                # Redirect back to the referring page
                return HttpResponseRedirect(referring_url)
            else:
                # If there's no referring URL, redirect to a default page
                return redirect('scanned-products', username=user.username)

        elif 'searched-products-btn' in request.POST:
            searched_product_ids = request.POST.getlist('searched-product-ids')
            for product_id in searched_product_ids:
                searched_product = CashierDynamicProducts.objects.get(id=int(product_id))
                scanned_product_list.product.add(searched_product)

                scanned_product_list.summed_product_price += searched_product.price
                scanned_product_list.save()

                searched_product.scanned_quantity += 1
                searched_product.quantity -= 1
                searched_product.save()

                main_product = Product.objects.get(
                    name=searched_product.name
                )
                main_product.quantity = searched_product.quantity
                main_product.save()

                dynamic_products = CashierDynamicProducts.objects.filter(
                    name=searched_product.name
                ).exclude(cashier=user)

                for dy_prod in dynamic_products:
                    dy_prod.quantity = searched_product.quantity
                    dy_prod.save()

            referring_url = request.META.get('HTTP_REFERER')

            if referring_url:
                # Redirect back to the referring page
                return HttpResponseRedirect(referring_url)
            else:
                # If there's no referring URL, redirect to a default page
                return redirect('scanned-products', username=user.username)

    return render(request, 'scanned_product_cashier.html', {
        'scanned_sum': scanned_sum,
        'scanned_products': scanned_products,
        'products': products,
        'user': user,
        'scanned_product_header': scanned_product_header,
        'add_quantity_form': add_quantity_form,
        'notifications': notifications,
        'user_belongs_to_admin': user_belongs_to_admin,
        'scanned_products_header': scanned_products_header,
        'searched_products': searched_products,
        'searched_product_name_input': searched_product_name_input,
    })


@has_expired
def async_product_search(request, username):
    user = User.objects.get(username=username)
    products = CashierDynamicProducts.objects.filter(cashier=user, quantity__gt=0)
    searched_product_name_input = request.POST.get('search-product-input')

    searched_products = None
    if searched_product_name_input:
        searched_products = CashierDynamicProducts.objects.filter(
            cashier=user,
            quantity__gt=0,
            name__icontains=searched_product_name_input
        )

    return render(request, 'product_search_results.html', {
        'searched_products': searched_products,
        'products': products,
        'searched_product_name_input': searched_product_name_input
    })


@has_expired
@login_required(login_url='login')
def update_scanned_products(request, username):
    cashier = User.objects.get(username=username)
    scanned_products = ScannedProducts.objects.filter(cashier=cashier)
    product_list = []

    for products in scanned_products:
        for product in products.product.all():
            product_data = {
                'name': product.name,
                'scanned_quantity': product.scanned_quantity,
                'price': product.scanned_quantity * product.price,
                'barcode': product.barcode,
                'id': product.id,
            }
            product_list.append(product_data)

    return JsonResponse({'products': product_list})


@has_expired
@allowed_users(allowed_roles=['cashier'])
@login_required(login_url='login')
def add_quantity_product(request, username, barcode):

    user = User.objects.get(username=username)

    scanned_product_list = ScannedProducts.objects.get(cashier=user)

    # scanned_product_barcode = request.GET.get['video']
    scanned_product = CashierDynamicProducts.objects.get(barcode=barcode, cashier=user)
    # scanned_product_list.product.add(scanned_product)

    scanned_product_list.summed_product_price += scanned_product.price
    scanned_product_list.save()

    scanned_product.scanned_quantity += 1
    # Newly Added Somehow I wanted to test this out, to make the quantities change in real time. 1/15/2024
    scanned_product.quantity -= 1
    scanned_product.save()

    referring_url = request.META.get('HTTP_REFERER')

    if referring_url:
        # Redirect back to the referring page
        return HttpResponseRedirect(referring_url)
    else:
        # If there's no referring URL, redirect to a default page
        return redirect('scanned-products', username=user.username)


@has_expired
@allowed_users(allowed_roles=['cashier'])
@login_required(login_url='login')
def subtract_quantity_product(request, username, barcode):
    user = User.objects.get(username=username)

    scanned_product_list = ScannedProducts.objects.get(cashier=user)

    # scanned_product_barcode = request.GET.get['video']
    scanned_product = CashierDynamicProducts.objects.get(barcode=barcode, cashier=user)
    # scanned_product_list.product.add(scanned_product)

    scanned_product_list.summed_product_price -= scanned_product.price
    scanned_product_list.save()

    scanned_product.scanned_quantity -= 1
    # Newly Added Somehow I wanted to test this out, to make the quantities change in real time. 1/25/2024
    # It works like heaven
    scanned_product.quantity += 1
    scanned_product.save()

    referring_url = request.META.get('HTTP_REFERER')

    if referring_url:
        # Redirect back to the referring page
        return HttpResponseRedirect(referring_url)
    else:
        # If there's no referring URL, redirect to a default page
        return redirect('scanned-products', username=user.username)


@has_expired
@allowed_users(allowed_roles=['cashier'])
@login_required(login_url='login')
def remove_scanned_product(request, username, barcode):
    user = User.objects.get(username=username)

    scanned_products_header = ScannedProductHeader.objects.get(cashier=user)
    scanned_products_header.name = None
    scanned_products_header.barcode = None
    scanned_products_header.price = None
    scanned_products_header.expiry_date = None
    scanned_products_header.save()

    scanned_product_list = ScannedProducts.objects.get(cashier=user)

    # scanned_product_barcode = request.GET.get['video']
    scanned_product = CashierDynamicProducts.objects.get(barcode=barcode, cashier=user)
    scanned_product.quantity += scanned_product.scanned_quantity

    main_products = Product.objects.filter(name=scanned_product.name)
    for main_product in main_products:
        main_product.quantity = scanned_product.quantity
        main_product.save()

    dynamic_products = CashierDynamicProducts.objects.filter(name=scanned_product.name)
    for dynamic_product in dynamic_products:
        dynamic_product.quantity = scanned_product.quantity
        dynamic_product.save()

    scanned_product_list.product.remove(scanned_product)

    scanned_product_list.summed_product_price -= (scanned_product.price * scanned_product.scanned_quantity)
    scanned_product_list.save()

    scanned_product.scanned_quantity = 0
    scanned_product.save()

    referring_url = request.META.get('HTTP_REFERER')

    if referring_url:
        # Redirect back to the referring page
        return HttpResponseRedirect(referring_url)
    else:
        # If there's no referring URL, redirect to a default page
        return redirect('scanned-products', username=user.username)


@has_expired
@allowed_users(allowed_roles=['cashier'])
@login_required(login_url='login')
def sell_scanned_products(request, username):
    user = User.objects.get(username=username)
    scanned_products = ScannedProducts.objects.filter(cashier=user)
    admin_user = User.objects.filter(groups__name='admin')
    # Let's retrieve the models that we need for this operation
    product_types = ProductType.objects.all()
    products = CashierDynamicProducts.objects.filter(cashier=user)
    cashier = ScannedProducts.objects.get(cashier=user)
    scanned_products_header = ScannedProductHeader.objects.get(cashier=user)

    for product in cashier.product.all():
        # Let's create a SoldProduct object from products that are sold right after they are scanned.
        sold_product = SoldProduct.objects.create(
            cashier=user,
            name=product.name,
            type=product.type,
            barcode=product.barcode,
            price=product.price,
            expiry_date=product.expiry_date,
            sold_quantity=product.scanned_quantity,
            date_sold=timezone.now(),
            date_sold_no_time=timezone.now(),
        )
        sold_product.save()

        # Let's create a replica of the product to avoid the sold products being erased when the original product
        # stock is depleted

        product_sold, created = SoldProductHub.objects.get_or_create(cashier=user)
        product_sold.product.add(sold_product)
        product_sold.save()

        # The scanned quantity of the scanned product to revert it to its original state.
        product.scanned_quantity = 0
        product.save()

    # Create and derive the products that are temporarily stored in this hub
    sold_products_hub = SoldProductHub.objects.get(cashier=user)
    product_sold = [product for product in sold_products_hub.product.all()]

    cash = request.POST.get('cash')

    # Let's create a SoldProducts object
    sold_products = SoldProducts.objects.create(
        cashier=user,
        total_price=cashier.summed_product_price,
        date_sold=timezone.now(),
        date_sold_no_time=timezone.now(),
        cash=int(cash),
        change=(int(cash) - cashier.summed_product_price),
    )
    sold_products.product.add(*product_sold)
    sold_products.save()

    # Sets the scanned product header to default
    scanned_products_header.name = None
    scanned_products_header.barcode = None
    scanned_products_header.price = None
    scanned_products_header.expiry_date = None
    scanned_products_header.change = sold_products.change
    scanned_products_header.save()

    # remove the products in the ScannedProducts Hub
    cashier.summed_product_price = 0
    product = [product for product in products]
    cashier.product.remove(*product)
    cashier.save()

    product_main = Product.objects.filter(quantity__lte=0)
    for product in product_main:
        sold_out_product = SoldOutProduct.objects.create(
            name=product.name,
            barcode=product.barcode,
            price=product.price,
            type=product.type,
        )
        for admin in admin_user:
            notification = Notification.objects.create(
                user=admin,
                title=f'Sold Out Product: {product.name}',
                message=f'This is to notify you that the product, "{product.name}" is now out of stock.'
            )

            notification.save()

        sold_out_product.save()
        product.delete()

    dynamic_products = CashierDynamicProducts.objects.filter(quantity__lte=0)
    for dynamic_product in dynamic_products:
        dynamic_product.delete()

    products = CashierDynamicProducts.objects.filter(cashier=user)
    product_types_list = {}
    for producto in products:
        product_type = producto.type
        if product_type not in product_types_list:
            product_types_list[product_type] = [producto]
        else:
            product_types_list[product_type].append(producto)

    # Add the SoldProduct objects from the SoldProductHub model
    sold_products_hub.product.remove(*product_sold)
    sold_products_hub.save()

    referring_url = request.META.get('HTTP_REFERER')

    if referring_url:
        # Redirect back to the referring page
        return HttpResponseRedirect(referring_url)

    return render(request, 'products_list.html', {
        'scanned_products_header': scanned_products_header,
        'products': products,
        'scanned_products': scanned_products,
        'product_types': product_types,
        'product_types_list': product_types_list,
    })


@has_expired
@login_required(login_url='login')
def add_dynamic_products(request, username):
    user = User.objects.get(username=username)
    products = Product.objects.all()
    for product in products:
        dynamic_cashier_products = CashierDynamicProducts.objects.create(
            cashier=user,
            name=product.name,
            barcode=product.barcode,
            price=product.price,
            expiry_date=product.expiry_date,
            quantity=product.quantity,
            scanned_quantity=product.scanned_quantity,
        )
        dynamic_cashier_products.save()
    return redirect('homepage')


@has_expired
@allowed_users(allowed_roles=['admin'])
@login_required(login_url='login')
def products_sold(request):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    notifications = Notification.objects.filter(removed=False)
    sold_products = SoldProducts.objects.all().order_by('-date_sold')
    return render(request, 'sold_products.html', {
        'sold_products': sold_products,
        'notifications': notifications,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
@allowed_users(allowed_roles=['admin'])
@login_required(login_url='login')
def expired_products(request):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    notifications = Notification.objects.filter(removed=False)
    current_date = timezone.now()
    thirty_days_ago = current_date - timedelta(days=30)

    expired_product = Product.objects.filter(expiry_date__lte=current_date, expiry_date__gte=thirty_days_ago)

    thirty_days_from_now = current_date + timedelta(days=30)

    upcoming_expiring_products = Product.objects.filter(expiry_date__gte=current_date,
                                                        expiry_date__lte=thirty_days_from_now)

    return render(request, 'expiration.html', {
        'expired_products': expired_product,
        'upcoming_expiring_products': upcoming_expiring_products,
        'notifications': notifications,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
@allowed_users(allowed_roles=['admin'])
@login_required(login_url='login')
def reports(request):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    notifications = Notification.objects.filter(removed=False)
    return render(request, 'reports.html', {
        'notifications': notifications,
        'user_belongs_to_admin': user_belongs_to_admin
    })


@has_expired
@allowed_users(allowed_roles=['admin'])
@login_required(login_url='login')
def sold_out_products(request):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    notifications = Notification.objects.filter(removed=False)
    products = SoldOutProduct.objects.all()
    return render(request, 'sold_out_products.html', {
        'products': products,
        'notifications': notifications,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
@allowed_users(allowed_roles=['admin'])
@login_required(login_url='login')
def select_products_sell(request, username):
    notifications = Notification.objects.filter(removed=False)
    cashier = User.objects.get(username=username)
    products = CashierDynamicProducts.objects.filter(cashier=cashier)
    scanned_products = ScannedProducts.objects.filter(cashier=cashier)
    # scanned_products_header = ScannedProductHeader.objects.get_or_create(cashier=cashier)
    scanned_products_header = ScannedProductHeader.objects.get(cashier=cashier)
    product_types = ProductType.objects.all()
    default_quantity = 1
    default_cash = 0

    scanned_product = ScannedProducts.objects.get(cashier=cashier)
    total_price_scanned_products = scanned_product.summed_product_price

    if request.method == 'POST':
        selected_products = request.POST.getlist('selectedProducts')
        selected_products_quantity = request.POST.getlist(f'product-quantity')
        scanned_product_list = ScannedProducts.objects.get(cashier=cashier)
        scanned_product_list.product.add(*selected_products)
        filtered_quantity = [value for value in selected_products_quantity if value]

        for id, quantity in zip(selected_products, filtered_quantity):
            product = CashierDynamicProducts.objects.get(id=id)
            product.scanned_quantity += int(quantity)
            scanned_product_list.summed_product_price += product.price * int(quantity)
            product.quantity -= int(quantity)
            product.save()

            main_products = Product.objects.filter(name=product.name)
            for main_product in main_products:
                main_product.quantity = product.quantity
                main_product.save()

            dynamic_products = CashierDynamicProducts.objects.filter(name=product.name)
            for dynamic_product in dynamic_products:
                dynamic_product.quantity = product.quantity
                dynamic_product.save()

        scanned_product_list.save()

    product_types_list = {}
    for producto in products:
        product_type = producto.type
        if product_type not in product_types_list:
            product_types_list[product_type] = [producto]
        else:
            product_types_list[product_type].append(producto)

    scanned_length = ScannedProducts.objects.filter(cashier=cashier).annotate(scanned_length=Count('product'))
    scanned_sum = scanned_length.aggregate(length_sum=Sum('scanned_length'))['length_sum']

    return render(request, 'select_sell_products.html', {
        'products': products,
        'default_quantity': default_quantity,
        'scanned_products': scanned_products,
        'product_types_list': product_types_list,
        'product_types': product_types,
        'notifications': notifications,
        'scanned_sum': scanned_sum,
        'default_cash': default_cash,
        'total_price_scanned_products': total_price_scanned_products,
        'scanned_products_header': scanned_products_header,
    })


@has_expired
@allowed_users(allowed_roles=['cashier'])
@login_required(login_url='login')
def select_products_sell_cashier_group(request, username):
    notifications = Notification.objects.filter(removed=False)
    cashier = User.objects.get(username=username)
    products = CashierDynamicProducts.objects.filter(cashier=cashier)
    scanned_products = ScannedProducts.objects.filter(cashier=cashier)
    scanned_products_header = ScannedProductHeader.objects.get(cashier=cashier)
    product_types = ProductType.objects.all()
    default_quantity = 1
    default_cash = 0

    scanned_product = ScannedProducts.objects.get(cashier=cashier)
    total_price_scanned_products = scanned_product.summed_product_price

    if request.method == 'POST':
        selected_products = request.POST.getlist('selectedProducts')
        selected_products_quantity = request.POST.getlist(f'product-quantity')
        scanned_product_list = ScannedProducts.objects.get(cashier=cashier)
        scanned_product_list.product.add(*selected_products)
        filtered_quantity = [value for value in selected_products_quantity if value]

        for id, quantity in zip(selected_products, filtered_quantity):
            product = CashierDynamicProducts.objects.get(id=id)
            product.scanned_quantity += int(quantity)
            scanned_product_list.summed_product_price += product.price * int(quantity)
            product.quantity -= int(quantity)
            product.save()

            main_products = Product.objects.filter(name=product.name)
            for main_product in main_products:
                main_product.quantity = product.quantity
                main_product.save()

            dynamic_products = CashierDynamicProducts.objects.filter(name=product.name)
            for dynamic_product in dynamic_products:
                dynamic_product.quantity = product.quantity
                dynamic_product.save()

        scanned_product_list.save()

    product_types_list = {}
    for producto in products:
        product_type = producto.type
        if product_type not in product_types_list:
            product_types_list[product_type] = [producto]
        else:
            product_types_list[product_type].append(producto)

    scanned_length = ScannedProducts.objects.filter(cashier=cashier).annotate(scanned_length=Count('product'))
    scanned_sum = scanned_length.aggregate(length_sum=Sum('scanned_length'))['length_sum']

    return render(request, 'select_sell_products_cashier.html', {
        'products': products,
        'default_quantity': default_quantity,
        'scanned_products': scanned_products,
        'product_types_list': product_types_list,
        'product_types': product_types,
        'notifications': notifications,
        'scanned_sum': scanned_sum,
        'default_cash': default_cash,
        'total_price_scanned_products': total_price_scanned_products,
        'scanned_products_header': scanned_products_header,
    })


@has_expired
@login_required(login_url='login')
def expired_products_json(request):
    admin_users = User.objects.filter(groups__name='admin')
    user = User.objects.get(username=request.user)
    current_date = timezone.now()
    thirty_days_ago = current_date - timedelta(days=30)
    now = timezone.now().date()
    # Format the date and time into the desired numeric representation

    expired_product = Product.objects.filter(expiry_date__lte=current_date, expiry_date__gte=thirty_days_ago)
    # for user in admin_users:
    for product in expired_product:
        expired_notifications, created = Notification.objects.get_or_create(
            user=user,
            title=f'Expired Product: {product.name}',
            message=f'This is to notify you that the product, "{product.name}" is pass the expiration date.',
            identifier=f'Expired Products Identifier: {product.name}-{product.id}'
        )

        expired_notifications.save()

    thirty_days_from_now = current_date + timedelta(days=30)

    upcoming_expiring_products = Product.objects.filter(expiry_date__gte=current_date,
                                                        expiry_date__lte=thirty_days_from_now)
    # for user in admin_users:
    for product in upcoming_expiring_products:
        difference = product.expiry_date - now
        days_until_expiry = difference.days
        if days_until_expiry == 1:
            days_label = 'day'
        else:
            days_label = 'days'
        upcoming_expiration, created = Notification.objects.get_or_create(
            user=user,
            title=f'Product Nearing Expiration: {product.name}',
            message=f'This is to notify you that the product, "{product.name}", '
                    f'will be expired {days_until_expiry} {days_label} from now.',
            identifier=f' Upcoming Expirations Identifier: {product.name}-{product.id}'
        )
        upcoming_expiration.save()

    notifications = Notification.objects.filter(removed=False, user__username=request.user)
    notification_list = []

    unseen_notifications = Notification.objects.filter(is_seen=False, user__username=request.user).annotate(
        unseen_notifications=Count('is_seen'))
    unseen_sum = unseen_notifications.aggregate(unseen=Sum('unseen_notifications'))['unseen']

    for notification in notifications:
        notification_data = {
            'title': notification.title,
            'message': notification.message,
            'time': naturaltime(notification.date_time),
            'id': notification.id,
            'unseen': unseen_sum
        }
        notification_list.append(notification_data)

    return JsonResponse({
        'notification': notification_list,
    })


@has_expired
@login_required(login_url='login')
def remove_notification(request, title, id):
    notification = Notification.objects.get(title=title, id=id)
    notification.removed = True
    notification.save()

    # Assuming you want to send a JSON response to confirm the removal
    return JsonResponse({'message': 'Notification removed successfully'})


@has_expired
@login_required(login_url='login')
def seen_notifications(request):
    authenticated_user = request.user
    user_object = User.objects.get(username=authenticated_user)
    try:
        notifications = Notification.objects.filter(user=user_object)
        for notification in notifications:
            notification.is_seen = True
            notification.save()

        # Return a success response
        return JsonResponse({'status': 'success'})

    except ValueError as e:
        # Print the error for debugging purposes
        print('Error marking notifications as seen:', str(e))

        # Return an error response
        return JsonResponse({'status': 'error', 'message': str(e)})


@has_expired
@allowed_users(allowed_roles=['cashier'])
@login_required(login_url='login')
def clear_scanned_header_change(request, username):
    cashier = User.objects.get(username=username)
    scanned_products_header = ScannedProductHeader.objects.get(cashier=cashier)

    scanned_products_header.change = None
    scanned_products_header.save()

    return render(request, 'clear_change.html', {
        'scanned_products_header': scanned_products_header,
    })


@has_expired
@allowed_users(allowed_roles=['encoder'])
@login_required(login_url='login')
def product_info(request, id):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    product = Product.objects.get(id=id)
    product_form = ChangeProductForm(instance=product)
    user_belongs_to_encoder = request.user.groups.filter(name='encoder').exists()
    if request.method == 'POST':
        form = ChangeProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            product_data = form.save(commit=False)
            dynamic_products = CashierDynamicProducts.objects.filter(Q(name=product_data.name) |
                                                                     Q(barcode=product_data.barcode))
            for dynamic_product in dynamic_products:
                dynamic_product.name = product_data.name
                dynamic_product.type = product_data.type
                dynamic_product.barcode = product_data.barcode
                dynamic_product.price = product_data.price
                dynamic_product.expiry_date = product_data.expiry_date
                dynamic_product.quantity = product_data.quantity
                dynamic_product.image = product_data.image
                dynamic_product.save()
            product_data.save()
            return redirect('product-info', id=product.id)

    return render(request, 'product_profile.html', {
        'product': product,
        'product_form': product_form,
        'user_belongs_to_encoder': user_belongs_to_encoder,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
@allowed_users(allowed_roles=['encoder'])
@login_required(login_url='login')
def products_all(request):
    user_belongs_to_admin = request.user.groups.filter(name='admin').exists()
    user_belongs_to_encoder = request.user.groups.filter(name='encoder').exists()
    notifications = Notification.objects.filter(removed=False)
    products = Product.objects.all()
    product_types_list = {}
    for producto in products:
        product_type = producto.type
        if product_type not in product_types_list:
            product_types_list[product_type] = [producto]
        else:
            product_types_list[product_type].append(producto)
    return render(request, 'products.html', {
        'product_types_list': product_types_list,
        'notifications': notifications,
        'user_belongs_to_encoder': user_belongs_to_encoder,
        'user_belongs_to_admin': user_belongs_to_admin,
    })


@has_expired
@login_required(login_url='login')
def add_expense(request):
    expense_description = request.POST.get('expense-description')
    expense_amount = request.POST.get('expense-amount')
    expense_document = request.FILES.get('expense-document')
    if request.method == 'POST':
        Expenses.objects.create(
            expense=float(expense_amount),
            description=expense_description,
            image=expense_document,
            date=timezone.now(),
            date_no_time=timezone.now().date()
        )
        # expense_object, created = Expenses.objects.get_or_create(
        #     date_no_time=timezone.now().date(),
        #     defaults={
        #         'date': timezone.now(),
        #     }
        # )
        # expense_object.expense = float(expense_amount)
        # expense_object.description = expense_description
        # expense_object.image = expense_document
        # expense_object.save()

        http_referrer = request.META.get('HTTP_REFERER')
        if http_referrer:
            return HttpResponseRedirect(http_referrer)
        else:
            return redirect('homepage')

    return render(request, 'add_expense.html', {

    })


@has_expired
def product_types(request):
    product_type_objects = ProductType.objects.filter(removed=False)
    if request.method == 'POST':
        if 'change-product-type-btn' in request.POST:
            product_type_name = request.POST.get('product-type')
            product_type_id = int(request.POST.get('product-type-id'))

            derived_product_type = ProductType.objects.get(id=product_type_id)
            derived_product_type.product_type = product_type_name
            derived_product_type.save()

    return render(request, 'product_type.html', {
        'product_type_objects': product_type_objects,
    })


@has_expired
def remove_product_type(request, name, id):
    product_type = ProductType.objects.get(product_type=name, id=id)
    product_type.removed = True
    product_type.save()

    http_referer = request.META.get('HTTP_REFERER')
    if http_referer:
        return HttpResponseRedirect(http_referer)
    else:
        return redirect('homepage')


@superuser
def expiry_date_conf_page(request):
    # device_user = os.getlogin()
    #
    # # Determine the appropriate command based on the OS
    # command = ''
    # if os.name == 'nt':  # Windows
    #     command = 'wmic bios get serialnumber'
    # elif os.name == 'posix':  # Unix-based systems
    #     if 'darwin' in os.sys.platform:  # MacOS
    #         command = "system_profiler SPHardwareDataType | grep 'Serial Number'"
    #     else:  # Linux
    #         command = 'cat /sys/class/dmi/id/product_serial'
    #
    # # Execute the command and retrieve the serial number
    # serial_number = subprocess.check_output(command, shell=True).decode().strip()
    # serial_number_lines = serial_number.splitlines()
    #
    # # Take the last line which should be the actual serial number
    # serial_number = serial_number_lines[-1].strip()

    try:
        device_user = os.getlogin()
    except OSError:
        device_user = getpass.getuser()

    serial_number = None

    if os.name == 'nt':  # Windows
        command = 'wmic bios get serialnumber'
    elif os.name == 'posix':  # Unix-based systems
        if 'darwin' in os.sys.platform:  # MacOS
            command = "system_profiler SPHardwareDataType | grep 'Serial Number'"
        else:  # Linux
            command = 'cat /sys/class/dmi/id/product_serial 2>/dev/null'

        try:
            serial_number = subprocess.check_output(command, shell=True).decode().strip()
            serial_number_lines = serial_number.splitlines()
            serial_number = serial_number_lines[-1].strip()
        except subprocess.CalledProcessError:
            # Handle the error, maybe log it or set a default/fallback serial number
            serial_number = "UNKNOWN_SERIAL"

    if serial_number is None:
        serial_number = "UNKNOWN_SERIAL"

    device, created = DeviceInformation.objects.get_or_create(
        name=f'{device_user}-{serial_number}',
        defaults={'registration_date': timezone.now().date()}
    )

    if request.method == 'POST':
        expiry_date_str = request.POST.get('expiry-date')
        expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d')
        device.expiry_date = expiry_date
        device.save()

        return redirect('homepage')

    return render(request, 'expiry_date_conf.html', {
        'device': device,
    })


def resubscribe(request):
    return render(request, 'resubscribe.html', {

    })


def refund_search(request):
    receipt_id = request.POST.get('receipt-id')
    if request.method == 'POST':
        if SoldProducts.objects.filter(id=receipt_id).exists():
            return redirect('refund', sold_products_id=receipt_id)
        else:
            messages.error(request, 'Receipt Does not Exists!')

    return render(request, 'refund-receipt-search.html', {

    })


def refund(request, sold_products_id):
    sold_products_obj = SoldProducts.objects.get(id=sold_products_id)
    if request.method == 'POST':
        refund_quantity = int(request.POST.get('refund-quantity'))
        product_id = int(request.POST.get('product-id'))
        derived_sold_products_id = int(request.POST.get('sold-refunded-product-id'))

        derived_sold_products_obj = SoldProducts.objects.get(id=derived_sold_products_id)

        revenue = Revenue.objects.get(date=derived_sold_products_obj.date_sold)

        refund_product = SoldProduct.objects.get(id=product_id)

        refund_product.refunded_quantity = refund_product.sold_quantity
        refund_product.sold_quantity -= refund_quantity
        refund_product.refunded = True
        refund_product.refund_date = timezone.now().date()
        refund_product.refund_date_time = timezone.now()

        derived_sold_products_obj.refunded_total_price = derived_sold_products_obj.total_price
        derived_sold_products_obj.refunded = True
        derived_sold_products_obj.total_price -= refund_quantity * refund_product.price
        derived_sold_products_obj.refunded_change = derived_sold_products_obj.change
        derived_sold_products_obj.change += refund_quantity * refund_product.price

        revenue.revenue -= refund_product.sold_quantity * refund_product.price

        derived_sold_products_obj.save()
        refund_product.save()
        revenue.save()

        http_referrer = request.META.get('HTTP_REFERER')
        if http_referrer:
            return HttpResponseRedirect(http_referrer)
        else:
            return redirect('homepage')

    return render(request, 'refund.html', {
        'sold_products_obj': sold_products_obj,
    })


def refunded_product(request, sold_product_id, sold_products_id):
    sold_product_obj = SoldProduct.objects.get(id=sold_product_id)
    sold_product_obj.refunded = True
    sold_product_obj.refunded_quantity = sold_product_obj.sold_quantity
    sold_product_obj.refunded_date = timezone.now().date()
    sold_product_obj.refunded_date_time = timezone.now()

    revenue = Revenue.objects.get(date=sold_product_obj.date_sold)
    revenue.revenue -= sold_product_obj.multiply()

    # sold_product_obj.sold_quantity = 0

    sold_products_obj = SoldProducts.objects.get(id=sold_products_id)
    sold_products_obj.refunded = True
    sold_products_obj.refunded_total_price = sold_products_obj.total_price
    sold_products_obj.total_price -= sold_product_obj.refunded_quantity * sold_product_obj.price
    sold_products_obj.refunded_change = float(sold_products_obj.change)
    sold_products_obj.change += float((sold_product_obj.refunded_quantity * sold_product_obj.price))

    sold_product_obj.save()
    sold_products_obj.save()
    revenue.save()

    referring_url = request.META.get('HTTP_REFERER')

    if referring_url:
        return HttpResponseRedirect(referring_url)
    else:
        return redirect('homepage')
